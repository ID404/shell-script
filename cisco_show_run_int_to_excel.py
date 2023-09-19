#此程序主要是将思科交换机 show run 信息中的接口信息整理成excel表格
import sys
import re
import openpyxl
from datetime import datetime

# 检查命令行参数
if len(sys.argv) < 2:
    print("请提供要分析的文件名作为命令行参数。")
    sys.exit(1)

# 获取要分析的文件名
filename = sys.argv[1]

# 创建一个新的Excel工作簿
workbook = openpyxl.Workbook()
sheet = workbook.active

# 打开文件并读取内容
with open(filename, 'r') as file:
    content = file.read()

# 使用正则表达式提取接口配置块
interface_regex = r"interface (\S+)(.*?)!"
interface_blocks = re.findall(interface_regex, content, re.DOTALL)

# 写入标题行
sheet.cell(row=1, column=1, value="interface")
sheet.cell(row=1, column=2, value="description")
sheet.cell(row=1, column=3, value="mode")
sheet.cell(row=1, column=4, value="vlan")
sheet.cell(row=1, column=5, value="channel-group")
sheet.cell(row=1, column=6, value="ip_address")

# 打印标题行
print("interface\t\t\tdescription\t\t\tmode\t\tvlan\t\tchannel-group\t\tip_address")

# 记录行号
row = 2

# 遍历接口配置块
for interface_block in interface_blocks:
    interface = interface_block[0]
    config = interface_block[1]

    # 使用正则表达式提取描述信息
    description_regex = r"description\s+(.+)"
    description = re.search(description_regex, config)
    description_value = description.group(1) if description else ""

    # 使用正则表达式提取switchport mode信息
    mode_regex = r"switchport mode (\S+)"
    mode = re.search(mode_regex, config)
    mode_value = mode.group(1) if mode else ""

    # 使用正则表达式提取vlan信息
    access_vlan_regex = r"switchport access vlan (\d+)"
    trunk_vlan_regex = r"switchport trunk allowed vlan (.+)"
    access_vlan = re.search(access_vlan_regex, config)
    trunk_vlan = re.search(trunk_vlan_regex, config)
    vlan_value = access_vlan.group(1) if access_vlan else (trunk_vlan.group(1) if trunk_vlan else "")

    # 使用正则表达式提取channel-group信息
    channel_group_regex = r"channel-group (\S+)"
    channel_group = re.search(channel_group_regex, config)
    channel_group_value = channel_group.group(1) if channel_group else ""

    # 使用正则表达式提取ip地址和掩码信息
    ip_regex = r"ip address (\S+) (\S+)"
    ip = re.search(ip_regex, config)
    ip_address_value = ip.group(1) if ip else ""
    subnet_mask_value = ip.group(2) if ip else ""

    # 将信息写入Excel表格
    sheet.cell(row=row, column=1, value=interface)
    sheet.cell(row=row, column=2, value=description_value)
    sheet.cell(row=row, column=3, value=mode_value)
    sheet.cell(row=row, column=4, value=vlan_value)
    sheet.cell(row=row, column=5, value=channel_group_value)
    sheet.cell(row=row, column=6, value=f"{ip_address_value}/{subnet_mask_value}")

    # 打印提取的信息
    print(f"{interface}\t\t{description_value}\t\t{mode_value}\t\t{vlan_value}\t\t{channel_group_value}\t\t{ip_address_value}/{subnet_mask_value}")

    # 增加行号
    row += 1

# 生成当前时间作为文件名的一部分
current_time = datetime.now().strftime("%Y%m%d%H%M%S")

# 定义保存的文件名
output_filename = f"output_{current_time}.xlsx"

# 保存Excel文件
workbook.save(output_filename)
print(f"信息已成功写入Excel文件：{output_filename}")

# 关闭文件
file.close()%
